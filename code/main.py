from code.keep_alive import keep_alive # type: ignore
import discord # type: ignore
from discord.ext import commands # type: ignore
import openpyxl # type: ignore
import os
import nest_asyncio # Importing nest_asyncio


# ---------------------------
# CONFIGURATION
# ---------------------------
BOT_TOKEN = "XXXXXXX"  # <-- Insert your bot token
EXCEL_FILE = "data.xlsx"
SHEET_NAME = "Sheet1"

# Map column names to their 1-based column index in the Excel sheet:
COLUMN_MAP = {
    "name": 1,
    "tag": 2,
    "rank": 3,
    "username": 4,
    "password": 5,
    "v?": 6,
    "email": 7,
    "sellable": 8
}

intents = discord.Intents.default()
intents.message_content = True
bot = commands.Bot(command_prefix="!", intents=intents)

wb = None
sheet = None

# ---------------------------
# BOT STARTUP
# ---------------------------
@bot.event
async def on_ready():
    print(f"Logged in as {bot.user} (ID: {bot.user.id})")
    load_excel_data()
    print("Excel data loaded.")
    print("Bot is ready to receive commands.")


def load_excel_data():
    """
    Loads the Excel file into global variables wb, sheet.
    Creates a new workbook if the file does not exist.
    """
    global wb, sheet
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if SHEET_NAME in wb.sheetnames:
            sheet = wb[SHEET_NAME]
        else:
            sheet = wb.active
    else:
        # If file doesn't exist, create a new one with the columns (A–H).
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = SHEET_NAME
        # Create header row
        sheet.append(["name", "tag", "rank", "username", "password", "v?", "email", "SELLABLE?"])
        wb.save(EXCEL_FILE)


def save_excel_data():
    """Saves the workbook to the Excel file."""
    if wb:
        wb.save(EXCEL_FILE)

# ---------------------------
# COMMANDS
# ---------------------------

@bot.command(name="search_rank")
async def search_rank(ctx, *, rank_query: str = ""):
    """
    Finds all rows where 'rank' (column C) exactly matches the given rank_query.
    Usage: !search_rank <rank>
    Example: !search_rank plat 1
    """
    if not sheet:
        await ctx.send("Excel data not loaded.")
        return

    rank_query_lower = rank_query.lower().strip()
    matches = []

    # Iterate rows from 2 to max (skip header)
    for row_idx in range(2, sheet.max_row + 1):
        rank_cell = sheet.cell(row=row_idx, column=COLUMN_MAP["rank"]).value
        if not rank_cell:
            continue

        if rank_cell.lower().strip() == rank_query_lower:
            # Build a response string
            row_data = []
            for col_name, col_index in COLUMN_MAP.items():
                cell_val = sheet.cell(row=row_idx, column=col_index).value
                row_data.append(f"{col_name}={cell_val}")

            row_str = f"**Row {row_idx}:** " + ", ".join(row_data)
            matches.append(row_str)

    if matches:
        response = "\n".join(matches)
        await ctx.send(f"**Accounts with rank '{rank_query}':**\n{response}")
    else:
        await ctx.send(f"No rows found with rank exactly matching '{rank_query}'.")


@bot.command(name="show_rank")
async def show_rank(ctx, *, name_query: str = ""):
    """
    Looks up a row by 'name' (column A) exactly and returns the rank.
    Usage: !show_rank <name>
    Example: !show_rank adi4386
    """
    if not sheet:
        await ctx.send("Excel data not loaded.")
        return

    name_query_lower = name_query.lower().strip()
    found = False

    for row_idx in range(2, sheet.max_row + 1):
        name_cell = sheet.cell(row=row_idx, column=COLUMN_MAP["name"]).value
        if not name_cell:
            continue

        if name_cell.lower().strip() == name_query_lower:
            rank_cell = sheet.cell(row=row_idx, column=COLUMN_MAP["rank"]).value
            if rank_cell:
                await ctx.send(f"The rank for '{name_cell}' is: **{rank_cell}**")
            else:
                await ctx.send(f"'{name_cell}' does not have a rank listed.")
            found = True
            break

    if not found:
        await ctx.send(f"No row found with name exactly matching '{name_query}'.")

@bot.command(name="update_name")
async def update_name(ctx, name_query: str, column_name: str, *, new_value: str):
    """
    Updates a cell in the row where the 'name' column (column A) matches name_query.
    Usage: !update_name <name> <column_name> <new_value>

    Example: !update_name adi4386 rank ascendant 3

    This finds the row whose 'name' == 'adi4386' (ignoring case),
    and updates the 'rank' column to 'ascendant 3'.
    """
    if not sheet:
        await ctx.send("Excel data not loaded.")
        return

    # Convert to lowercase for case-insensitive matching
    name_query_lower = name_query.lower()
    column_key = column_name.lower()

    # Validate column name
    if column_key not in COLUMN_MAP:
        valid_cols = ", ".join(COLUMN_MAP.keys())
        await ctx.send(f"Invalid column '{column_name}'. Valid columns: {valid_cols}")
        return

    found_row = None

    # Find the row where 'name' column matches name_query
    for row_idx in range(2, sheet.max_row + 1):
        name_cell_val = sheet.cell(row=row_idx, column=COLUMN_MAP["name"]).value
        if name_cell_val is None:
            continue

        # Compare ignoring case
        if name_cell_val.lower() == name_query_lower:
            found_row = row_idx
            break

    if not found_row:
        await ctx.send(f"No row found with name exactly matching '{name_query}'.")
        return

    # Update the specific column in the found row
    col_index = COLUMN_MAP[column_key]
    sheet.cell(row=found_row, column=col_index).value = new_value
    save_excel_data()

    await ctx.send(
        f"Updated row {found_row} where name='{name_query}', column '{column_name}' to '{new_value}'."
    )

@bot.command(name="update_cell")
async def update_cell(ctx, row_number: int, column_name: str, *, new_value: str):
    """
    Updates a specific cell in the given row number by column name.
    Usage: !update_cell <row_number> <column_name> <new_value>
    Example: !update_cell 2 rank ascendant 3
    - This updates row 2, 'rank' column to 'ascendant 3'.
    """
    if not sheet:
        await ctx.send("Excel data not loaded.")
        return

    # Validate the column name
    column_name_lower = column_name.lower()
    if column_name_lower not in COLUMN_MAP:
        valid_cols = ", ".join(COLUMN_MAP.keys())
        await ctx.send(f"Invalid column '{column_name}'. Valid columns: {valid_cols}")
        return

    # Check row range (first data row is 2)
    if row_number < 2 or row_number > sheet.max_row:
        await ctx.send(f"Row number {row_number} is out of range (2 - {sheet.max_row}).")
        return

    col_index = COLUMN_MAP[column_name_lower]
    # Update the cell value
    sheet.cell(row=row_number, column=col_index).value = new_value
    save_excel_data()

    await ctx.send(f"Updated row {row_number}, column '{column_name}' to '{new_value}'.")

# ---------------------------
# RUN THE BOT
# ---------------------------
keep_alive()
nest_asyncio.apply() # Applying nest_asyncio to allow nested event loops
bot.run(BOT_TOKEN)
